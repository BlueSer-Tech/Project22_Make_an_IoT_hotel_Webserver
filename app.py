#from imutils.video import VideoStream
import socket
from flask import Flask, render_template, Response, request, jsonify, redirect, url_for, flash, session
import cv2
from pyzbar import pyzbar
from excel_index import query_data_excel
import imutils
from project_function import *
from openpyxl import load_workbook
import json
import face_recognition
import os
import numpy as np
from send_notification_mail import send_code
def ipadd():
    hostname=socket.gethostname()
    IPAddr=socket.gethostbyname(hostname)
    return IPAddr
path3 = r'C:\Users\Huy\Documents\DATH\pile.xlsx'
path2 = r'C:\Users\Huy\Documents\DATH\main_data.xlsx'
path1 = r'C:\Users\Huy\Documents\DATH\form_register.xlsx'
string = '062201002945|233306838|Trần Đăng Huy|17112001|Nam|68/31 Hàm Nghi Tổ 8, Duy Tân, Thành phố Kon Tum, Kon Tum|22112021'
app = Flask(__name__)
app.secret_key = 'ab156cdth'
vc = cv2.VideoCapture(0)
process_this_frame = True
text= '' #Bien tam de lay du lieu khi quet ma QR
data=[]# Bien tam de lay du lieu tu text sau khi xu ly
data1= []
room_value_send = time_value_send = "none"
room_value = time_value = ""
turn_back = 0#Trả về trang quét mã QR khi admin xác nhận thành công
count = 0 #Trả về trang quét mã QR khi bị admin từ chối 3 lần
directory = r'C:\Users\Huy\Documents\DATH\face_detect'
training_list = []
folder_name= ''
img_counter = 0
time_counter = 1
face_locations = ''
known_face_encodings = []
xac_nhan = '' # Xac nhan chan dung
xac_nhan_khuon_mat = 0 # Nhan dang duoc khuon mat
email= ""
face_names = []
known_face_names = []
phone = ""
custom_request = ""
birth = ""
name = ""
cccd = ""
gender = ""
custom_birth= ""
custom_gender= ""
custom_room_id = ""
custom_check_in = ""
custom_check_out = ""
custom_cccd = ""
custom_name = ""
custom_phone = ""
custom_email = ""
custom_leng_stay = ""
custom_pile_money = 0
#Trang chính
@app.route('/', methods=['POST', 'GET'])
def main_page():
    return render_template("main_page.html")
#Trang đăng ký thông tin------------------------------------------------------------------------------------------
@app.route('/register', methods=['POST', 'GET'])
def register():
    return render_template("register_information.html")
@app.route('/register/choose_room', methods=['POST', 'GET'])
def choose_room():
    global path1, phone, email, birth, name, cccd, gender
    register_list = []
    if request.method == "POST":
        phone = request.form.get("phone")
        email = request.form.get("email")
        birth = request.form.get("birth")
        name = request.form.get("name")
        cccd = request.form.get("cccd")
        gender = request.form.get("gender")
        register_list.append(cccd)
        register_list.append(name)
        register_list.append(birth)
        register_list.append(gender)
        register_list.append(email)
        register_list.append(phone)
        query_data_excel(register_list, path1)
    return render_template("choose_room.html")
#Trang chọn thời gian và tính tiền--------------------------------------------------------------------------------
@app.route('/register/choose_room/<room>', methods=['POST', 'GET'])
def get_room(room):
    time_date = ""
    if room == 'room1':
        if request.method == "POST":
            time_date = request.form.get("date")
            print(time_date)
        return render_template('room1_registration.html', data = time_date)
    if room == 'room2':
        return render_template('room2_registration.html')
@app.route('/register/choose_room/<room>/calendar', methods=['POST', 'GET'])
def calendar(room):
    if room == 'room1':
        return render_template('calendar.html')
@app.route('/register/choose_room/<room>/pile', methods=['POST', 'GET'])
def pile(room):
    global path1, phone, email, birth, name, cccd, gender
    start_day = ""
    end_day = ""
    leng_stay = ""
    room_id = ""
    total_money = 0
    if room == 'room1':
        if request.method == "POST":
            start_day = request.form.get("start_day")
            leng_stay = request.form.get("length_stay")
            day, month, year = format_date(start_day)
            start_day = day + " " + month + " " + year
            end_day = calendar_operator(day, month, year, leng_stay)
            room_id = request.form.get("room_id")
            total_money = 200000 * int(leng_stay)
    return render_template('pile.html', start_day = start_day, end_day = end_day, leng_stay = leng_stay, total_money = total_money
                           , phone = phone, email = email, birth = birth, name = name, cccd = cccd, gender = gender,
                           room_id = room_id )
#Trang lấy mẫu khuôn mặt------------------------------------------------------------------------------------------
# @app.route('/face_detect', methods=['POST', 'GET'])
# def face_detect():
#     global img_counter
#     if request.method == "POST":
#         print(img_counter)
#         if img_counter >= 3:
#             return jsonify({"status": img_counter})
#     return render_template('face_detect.html')
# def gen1():
#     global img_counter, folder_name, directory
#     img_counter = 0
#     folder_name = "User{}".format(count_dir(directory) + 1)
#     make_folder(folder_name)
#     directory1 = r'C:\Users\Huy\Documents\DATH\face_detect\{}'.format(folder_name)
#     while True:
#         rval, frame1 = vc.read()
#         if process_this_frame:
#             frame = cv2.resize(frame1, (0, 0), fx=0.25, fy=0.25)
#             frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
#             # Resize frame of video to 1/4 size for faster face recognition processing
#             # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
#               # [:, :, ::-1]
#             # Find all the faces and face encodings in the current frame of video
#             face_locations = face_recognition.face_locations(frame)
#             #if(face_locations):
#
#             # face_encodings = face_recognition.face_encodings(frame, face_locations)
#             # if face_encodings != []:
#             #     face_encodings = np.array(face_encodings)
#             #     face_encodings = face_encodings.tolist()
#             #     known_face_encodings.append(face_encodings[0])
#             #     xac_nhan_khuon_mat = 1
#             if(face_locations and img_counter <= 3):
#                 img_name = "{}.png".format(img_counter)
#                 img_counter += 1
#                 os.chdir(directory1)
#                 cv2.imwrite(img_name, frame)
#             # if face_encodings != [] and time_counter % 10 == 0:
#             #     img_name = "opencv_frame_{}.png".format(img_counter)
#             #     os.chdir(directory)
#             #     cv2.imwrite(img_name, frame)
#             #     print("{} written!".format(img_name))
#             #     img_counter += 1
#             # time_counter += 1
#             # print(time_counter)
#             cv2.imwrite('t.jpg', frame1)
#
#             yield (b'--frame\r\n'
#                    b'Content-Type: image/jpeg\r\n\r\n' + open('t.jpg', 'rb').read() + b'\r\n')
# @app.route('/video_feed1')
# def video_feed1():
#     return Response(gen1(),
#                     mimetype='multipart/x-mixed-replace; boundary=frame')
#
# #Trang để training data------------------------------------------------------------------------------------------
# @app.route('/training', methods=['POST', 'GET'])
# def training():
#     global folder_name, training_list
#     if request.method == "POST":
#         directory1 = r'C:\Users\Huy\Documents\DATH\face_detect\{}'.format(folder_name)
#         NumberOfFile = count_file(directory1) - 1
#         for i in range(0, NumberOfFile):
#             os.chdir(directory1)
#             image = face_recognition.load_image_file("{}.png".format(i))
#             face_re = face_recognition.face_encodings(image)[0]
#             training_list.append(face_re)
#         return jsonify({"status": True})
#     return render_template('training.html')

#Trang để quét mã QR------------------------------------------------------------------------------------------
@app.route('/qr_detect', methods=['POST', 'GET'])
def qr_detect():
    global data, text
    if request.method == "POST":
        if text != "":
            data = step_list(text)
            # return redirect('/display_data')
        return jsonify({"status": data})
    return render_template('qr_detect.html')
def gen2():
    """Video streaming generator function."""
    global text, process_this_frame, directory, img_counter, time_counter, directory1, known_face_encodings, xac_nhan_khuon_mat
    while True:
        rval, frame = vc.read()
        barcodes = pyzbar.decode(frame)
        for barcode in barcodes:
            (x, y, w, h) = barcode.rect
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
            barcodeData = barcode.data.decode("utf-8")
            text = "{}".format(barcodeData)
            cv2.putText(frame, text, (x, y - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
        directory1 = r'C:\Users\Huy\Documents\DATH'
        os.chdir(directory1)
        cv2.imwrite('t.jpg', frame)
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + open('t.jpg', 'rb').read() + b'\r\n')
@app.route('/video_feed2')
def video_feed2():
    return Response(gen2(),
                    mimetype='multipart/x-mixed-replace; boundary=frame')

#Trang kiểm tra khuôn mặt vừa training với khuôn mặt ở trên CCCD-----------------------------------------------
# @app.route('/check_face', methods=['POST', 'GET'])
# def check_face():
#     if request.method == "POST":
#         return jsonify({"status":xac_nhan})
#     return render_template('check_face.html')
# def gen3():
#     global process_this_frame, xac_nhan, directory1, known_face_encodings, training_list, face_names, known_face_names, data
#     known_face_names.append(data[2])
#     while True:
#         rval, frame = vc.read()
#         frame1 = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
#         frame1 = cv2.cvtColor(frame1, cv2.COLOR_BGR2RGB)
#         face_locations = face_recognition.face_locations(frame1)
#         face_encodings = face_recognition.face_encodings(frame1, face_locations)
#         for face_encoding in face_encodings:
#                 # See if the face is a match for the known face(s)
#             matches = face_recognition.compare_faces(training_list, face_encoding)
#             name = "Unknown"
#             if True in matches:
#                 xac_nhan = 'True'
#                 first_match_index = matches.index(True)
#                 name = known_face_names[first_match_index]
#             face_names.append(name)
#             for (top, right, bottom, left), name in zip(face_locations, face_names):
#                     # Scale back up face locations since the frame we detected in was scaled to 1/4 size
#                 top *= 4
#                 right *= 4
#                 bottom *= 4
#                 left *= 4
#                     # Draw a box around the face
#                 cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
#                     # Draw a label with a name below the face
#                 cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
#                 font = cv2.FONT_HERSHEY_DUPLEX
#                 cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)                #face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
#                 #best_match_index = np.argmin(face_distances)
#
#         os.chdir(directory1)
#         cv2.imwrite('t.jpg', frame)
#         yield (b'--frame\r\n'
#                b'Content-Type: image/jpeg\r\n\r\n' + open('t.jpg', 'rb').read() + b'\r\n')
#
# @app.route('/video_feed3')
# def video_feed3():
#     return Response(gen3(),
#                     mimetype='multipart/x-mixed-replace; boundary=frame')

#Trang hiển thị thông tin------------------------------------------------------------------------------------------
@app.route('/display_data', methods=['POST', 'GET'])
def index1():
    global data, text, xac_nhan, known_face_encodings
    known_face_encodings = []
    xac_nhan= ''
    data = step_list(text)
    if request.method == "POST":
        return jsonify({'cccd': data[0], 'name': data[2], 'born':data[3], 'gender': data[4]})
    return render_template('display_data.html')
#Trang cảm ơn------------------------------------------------------------------------------------------
@app.route('/hienthi', methods=['POST', 'GET'])
def camon():
    global data, data1, email, text
    text = ""
    data1 = data.copy()
    if request.method == "POST":
        std = request.form.get("std")
        email = request.form.get("email")
        data1.append(std)
        data1.append(email)

        # if std in data:
        #     None
        # else:
        #     if data[-1].isnumeric() == False:
        #         data.append(std)
        #     else:
        #         data[-1]= std
    return render_template('camon.html')
#Trang chọn phòng----------------------------------------------------------------
@app.route('/chonphong', methods=['POST', 'GET'])
def chonphong():
    status_room1 = status_room2 = "None"
    wb = load_workbook('C:/Users/Huy/Documents/Doan.xlsx')
    ws = wb.active
    for i in range(2, ws.max_row + 1):
        row_status = [cell.value for cell in ws[i][2:11]]
        if 'room1' in row_status:
            status_room1 = 'exist'
        elif 'room2' in row_status:
            status_room2 = 'exist'
    if request.method == "POST":
        return jsonify({'status1': status_room1, 'status2':status_room2})
    return render_template('chonphong.html')
#Trang chọn loại phòng----------------------------------------------------------------
@app.route('/chonphong/<tien>', methods=['POST', 'GET'])
def thanhtoan(tien):
    if tien == 'room1_3h' or tien == 'room1_12h' or tien == 'room1_overnight':
        return render_template('room1.html')
    elif tien == 'room2_3h' or tien == 'room2_12h' or tien == 'room2_overnight':
        return render_template('room2.html')
@app.route('/admin', methods=['POST', 'GET'])
#Trang admin----------------------------------------------------------------
def admin():
    global time_value_send, room_value_send, data, data1, room_value, path2, path3,custom_check_out, custom_room_id, custom_request, time_value, turn_back, count, email, path2, custom_name, custom_cccd, custom_birth, custom_gender, custom_email, custom_phone, custom_check_in, custom_leng_stay, custom_pile_money
    room1 = room2 = [None]*11
    wb = load_workbook(path2)
    wk = load_workbook(path3)
    ws = wb.active
    wd = wk.active
    for i in range(2, ws.max_row + 1):
        row_status = [cell.value for cell in ws[i][0:11]]
        if 'room1' in row_status:
            room1= row_status
        elif 'room2' in row_status:
            room2= row_status


    if request.method == "POST":
        #nhận data
        so_phong = json.loads(request.data).get('room')
        thoi_gian= json.loads(request.data).get('time')
        verify = json.loads(request.data).get('verify')
        print("custom request: ",custom_request)
        print("Phong dat: ",room_value_send)
        print("Xac nhan: ",verify)
        if verify == "COMPLETE":
            turn_back = 1
            temp_list = list_insert(custom_check_in,custom_check_out, custom_leng_stay,custom_cccd, custom_name, custom_birth, custom_gender, custom_email, custom_phone, custom_pile_money, custom_room_id)
            print(temp_list)
            query_data_excel(temp_list, path3)
            custom_request = "NO"
        #xét data
        if room_value_send != "none" and room_value_send != "None" and room_value_send != None :
            room_value = room_value_send
            time_value = time_value_send
            room_value_send = "none"
            count += 1
        else:
            room_value = "none"
            time_value = "none"
        if so_phong != 'none' and thoi_gian!= 'none' and so_phong != '' and thoi_gian!= '' and so_phong != 'refuse':
            data1.append(so_phong)
            data1.append(thoi_gian)
            query_data_excel(data1, path2)#Save data to the database
            send_code(email)#Send the mail that contains 6-digits number to get into the room
            turn_back = 1
            count = 1
        elif so_phong == 'refuse':
            room_value_send= time_value_send = "none"
        #gửi data
        return jsonify({'cccd1': room1[1], 'name1': room1[3],'room1':room1[-1],
                        'cccd2': room2[1], 'name2': room2[3], 'room2':room2[-1], 'custom_request': custom_request,
                        'pile_cccd': custom_cccd,
                        'room': room_value, 'time':time_value, "pile_room" : custom_room_id, "pile_start_day": custom_check_in,
                        "pile_end_day": custom_check_out, "pile_leng_stay": custom_leng_stay, "pile_total_money": str(custom_pile_money)})

    return render_template('admin.html')
@app.route('/phong_cho', methods=['POST', 'GET'])
def xac_nhan():
    global time_value_send,room_value_send ,turn_back, count,custom_request,custom_name,custom_room_id, custom_cccd, custom_birth, custom_gender, custom_email, custom_phone, custom_check_in, custom_check_out, custom_leng_stay, custom_pile_money
    if request.method == 'POST':
        room_value_send = request.form.get("room")
        time_value_send = request.form.get("time")
        custom_request = request.form.get("request")

        custom_name = request.form.get("fname")
        custom_cccd = request.form.get("cccd")
        custom_birth = request.form.get("birth")
        custom_gender = request.form.get("gender")
        custom_email = request.form.get("email")
        custom_phone = request.form.get("phone")
        custom_room_id = request.form.get("room_id")
        custom_check_in = request.form.get("start_day")
        custom_check_out = request.form.get("end_day")
        custom_leng_stay = request.form.get("leng_stay")
        custom_pile_money = request.form.get("total_money")
        print(custom_leng_stay)
        if turn_back == 1 or count == 4:
            turn_back = 0
            count = 0
            return jsonify({"status": "turnback"})
    return render_template('phong_cho.html')

if __name__ == '__main__':
    app.run(host=f'{ipadd()}', debug=True, threaded=True)