from openpyxl import  load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from project_function import step_list
def query_data_excel(list, path):
    list.insert(0, datetime.now())
    wb = load_workbook(path)
    ws = wb.active
    #Lay du lieu trong database(excel)
    ws.append(list)
    wb.save(path)
    #------------------------------------
    '''elif(check_exist1(data, lst) == None):
        for i in range(len(lst1)):
            if lst[i] == data:
                lst2 = final_list(data, lst1[i][0], lst1[i][1])
                if lst2 != None:
                    for col in range(0,2):
                        char = get_column_letter(col+1)
                        ws[char + str(i+1)] = lst2[col]
                    wb.save('C:/Users/Huy/Documents/Doan.xlsx')'''

def delete_row(path):
    wb = load_workbook(path)
    ws = wb.active
    ws.delete_rows(2,1000)
    wb.save(path)

path3 = r'C:\Users\Huy\Documents\DATH\pile.xlsx'
path2 = r'C:\Users\Huy\Documents\DATH\main_data.xlsx'
path1 = r'C:\Users\Huy\Documents\DATH\form_register.xlsx'
delete_row(path3)
